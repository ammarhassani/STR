"""Retrospective import: 20 years of Excel history into STR.

Run: python tests_retrospective_import.py

Design: docs/superpowers/specs/2026-07-19-retrospective-import-design.md
"""
import os
import sys
import sqlite3
import tempfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

_fail = 0


def check(label, cond, detail=""):
    global _fail
    ok = bool(cond)
    print(("PASS " if ok else "FAIL ") + label + ("" if ok else f"  -- {detail}"))
    if not ok:
        _fail += 1


def _setup():
    from database.init_db import initialize_database
    from database.migrations import migrate_database
    from database.db_manager import DatabaseManager
    from services.logging_service import LoggingService
    from services.auth_service import AuthService
    from services.report_service import ReportService
    from services.report_number_service import ReportNumberService
    from services.activity_service import ActivityService
    from services.version_service import VersionService
    from services.security_service import SecurityService
    from services.retrospective_import import RetrospectiveImportService

    d = tempfile.mkdtemp()
    db = os.path.join(d, "r.db")
    initialize_database(db)
    migrate_database(db)
    conn = sqlite3.connect(db)
    conn.execute("UPDATE users SET password=?, must_change_password=0 WHERE username='admin'",
                 (SecurityService.hash_password('Admin@1234'),))
    for u, r in (('agent1', 'agent'), ('sup1', 'supervisor')):
        conn.execute("INSERT OR IGNORE INTO users (username,password,full_name,role,is_active,"
                     "created_by,must_change_password) VALUES (?,?,?,?,1,'SYSTEM',0)",
                     (u, SecurityService.hash_password('Pass@123'), u.upper(), r))
    conn.commit()
    conn.close()

    dbm = DatabaseManager(db)
    log = LoggingService(dbm, None, db_logging=False)
    auth = AuthService(dbm, log)
    log.set_auth_service(auth)
    act = ActivityService(dbm, log, auth)
    nums = ReportNumberService(dbm, log, auth)
    reports = ReportService(dbm, log, auth, act)
    vers = VersionService(dbm, log, auth, reports, act)
    reports.set_activity_service(act)
    reports.set_report_number_service(nums)
    reports.set_version_service(vers)
    imp = RetrospectiveImportService(dbm, log, auth, reports, nums)
    return d, dbm, auth, reports, nums, imp


def _rows(n=3, start=1, month="2016/04", cic_base=9100000000000000):
    """n clean historical rows, as the cleansed Excel would carry them."""
    return [{
        'report_number': f"{month}/{start + i:03d}",
        'sn': 10000 + start + i,
        'report_date': f"{(i % 28) + 1:02d}/04/2016",
        'reported_entity_name': f"Historic Entity {start + i}",
        'cic': str(cic_base + start + i),
        'report_classification': 'Money Laundering',
        'total_transaction': '250000',
    } for i in range(n)]


def _write_xlsx(path, rows, headers=None):
    from openpyxl import Workbook
    from services.retrospective_import import RetrospectiveImportService as R
    headers = headers or R.TEMPLATE_COLUMNS
    wb = Workbook()
    ws = wb.active
    ws.append(list(headers))
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    wb.save(path)
    return path


# ------------------------------------------------------------------- template
def test_template_is_generated_by_the_app():
    """The file always originates here, so the columns can never drift."""
    d, dbm, auth, reports, nums, imp = _setup()
    auth.authenticate('admin', 'Admin@1234')
    path = os.path.join(d, "template.xlsx")
    ok, msg = imp.write_template(path)
    check("admin can generate the blank template", ok, msg)
    check("the template file exists", os.path.exists(path))

    from openpyxl import load_workbook
    ws = load_workbook(path).active
    header = [c.value for c in ws[1]]
    check("header matches exactly what the importer reads",
          header == list(imp.TEMPLATE_COLUMNS), header)
    check("the four required fields are in the template",
          all(c in header for c in ('report_number', 'sn', 'report_date',
                                    'reported_entity_name')), header)
    check("the template has no data rows", ws.max_row == 1, ws.max_row)

    auth.authenticate('agent1', 'Pass@123')
    ok2, msg2 = imp.write_template(os.path.join(d, "nope.xlsx"))
    check("a non-admin cannot generate the template", not ok2, msg2)


# --------------------------------------------------------------- happy import
def test_clean_file_imports():
    d, dbm, auth, reports, nums, imp = _setup()
    auth.authenticate('admin', 'Admin@1234')
    f = _write_xlsx(os.path.join(d, "hist.xlsx"), _rows(3))

    ok, result = imp.import_file(f)
    check("a clean file imports", ok, result)
    check("every row landed", result.get('imported') == 3, result)

    rows = dbm.execute_with_retry(
        "SELECT report_number, sn, approval_status, reported_entity_name "
        "FROM reports ORDER BY report_number")
    check("three reports exist", len(rows) == 3, len(rows))
    check("the original report number is kept exactly",
          rows[0][0] == '2016/04/001', rows[0][0])
    check("the original serial is kept", rows[0][1] == 10001, rows[0][1])
    check("they land as archived, not approved",
          all(r[2] == 'archived' for r in rows), [r[2] for r in rows])


def test_numbers_are_registered_as_used():
    """'2016/04/555 is registered because it has been reserved by an archived
    report' -- the sequence must never re-issue it."""
    d, dbm, auth, reports, nums, imp = _setup()
    auth.authenticate('admin', 'Admin@1234')
    f = _write_xlsx(os.path.join(d, "h.xlsx"),
                    [dict(_rows(1)[0], report_number='2016/04/555', sn=20555)])
    ok, result = imp.import_file(f)
    check("import ok", ok, result)

    row = dbm.execute_with_retry(
        "SELECT owned_by, status, used_by_report_id FROM reserved_numbers "
        "WHERE report_number = '2016/04/555'")
    check("the number is registered in the numbering table", bool(row), row)
    check("it is marked used, not available", row and row[0][1] == 'used', row)
    check("it points at the archived report", row and row[0][2] is not None, row)

    # and the live sequence for that month must now start above it
    nums._now = lambda: __import__('datetime').datetime(2016, 4, 20)
    ok2, block, msg = nums.reserve_block('agent1', 1)
    check("a live reservation for that month skips the archived number",
          ok2 and block[0] == '2016/04/556', (block, msg))


# ------------------------------------------------------------- rejected files
def test_whole_file_is_rejected_and_nothing_is_written():
    d, dbm, auth, reports, nums, imp = _setup()
    auth.authenticate('admin', 'Admin@1234')
    rows = _rows(4)
    rows[2]['cic'] = '123'                       # not 16 digits
    rows[3]['report_date'] = 'March 2011'        # unparseable
    f = _write_xlsx(os.path.join(d, "bad.xlsx"), rows)

    ok, result = imp.import_file(f)
    check("a file with bad rows is refused", not ok, result)
    n = dbm.execute_with_retry("SELECT COUNT(*) FROM reports")[0][0]
    check("NOTHING was written -- not even the good rows", n == 0, n)
    nres = dbm.execute_with_retry("SELECT COUNT(*) FROM reserved_numbers")[0][0]
    check("no numbers were consumed either", nres == 0, nres)


def test_every_problem_is_reported_in_one_pass():
    """With 60k rows you cannot fix one error per upload."""
    d, dbm, auth, reports, nums, imp = _setup()
    auth.authenticate('admin', 'Admin@1234')
    rows = _rows(6)
    rows[0]['cic'] = '1' * 17          # too long to be a CIC
    rows[1]['cic'] = 'ABCDEFGHIJKLMNOP'
    rows[2]['report_date'] = 'nonsense'
    rows[3]['reported_entity_name'] = ''
    rows[4]['sn'] = ''
    f = _write_xlsx(os.path.join(d, "many.xlsx"), rows)

    ok, result = imp.import_file(f)
    check("refused", not ok)
    problems = result.get('problems') or {}
    check("all five bad rows are reported at once",
          result.get('bad_rows') == 5, result.get('bad_rows'))
    check("problems are grouped by type, not listed row by row",
          len(problems) >= 3, list(problems))
    for kind, info in problems.items():
        check(f"  '{kind}' carries a count and example rows",
              info.get('count') and info.get('rows'), info)


def test_short_cics_are_padded_not_rejected():
    """The bank prints a CIC without leading zeros and the report form pads it
    on entry. Historical rows carry the short form, so an imported CIC must be
    padded the same way or it would never match one typed today."""
    d, dbm, auth, reports, nums, imp = _setup()
    auth.authenticate('admin', 'Admin@1234')
    rows = _rows(1)
    rows[0]['cic'] = '12345'
    f = _write_xlsx(os.path.join(d, "pad.xlsx"), rows)
    ok, result = imp.import_file(f)
    check("a short numeric CIC does not reject the file", ok, result)
    stored = dbm.execute_with_retry("SELECT cic FROM reports")[0][0]
    check("it is padded to 16 digits", stored == '0000000000012345', stored)

    from services.intelligence_service import IntelligenceService
    hist = IntelligenceService(dbm, None).cic_history('0000000000012345')
    check("so a CIC typed today finds the historical filing", hist['count'] == 1, hist)


def test_duplicate_numbers_are_caught_inside_the_file():
    """Two rows claiming the same number -- a DB check alone would miss it."""
    d, dbm, auth, reports, nums, imp = _setup()
    auth.authenticate('admin', 'Admin@1234')
    rows = _rows(3)
    rows[2]['report_number'] = rows[0]['report_number']
    f = _write_xlsx(os.path.join(d, "dup.xlsx"), rows)
    ok, result = imp.import_file(f)
    check("a duplicate inside the file is caught", not ok, result)
    check("and named as such",
          any('duplicate' in k.lower() for k in (result.get('problems') or {})),
          list(result.get('problems') or {}))


def test_a_number_already_in_str_is_caught():
    d, dbm, auth, reports, nums, imp = _setup()
    auth.authenticate('admin', 'Admin@1234')
    f = _write_xlsx(os.path.join(d, "a.xlsx"), _rows(2))
    ok, _ = imp.import_file(f)
    check("first import ok", ok)
    ok2, result = imp.import_file(f)                    # same file again
    check("re-uploading the same file is refused", not ok2, result)
    n = dbm.execute_with_retry("SELECT COUNT(*) FROM reports")[0][0]
    check("and nothing is duplicated", n == 2, n)


def test_only_admin_may_import():
    d, dbm, auth, reports, nums, imp = _setup()
    auth.authenticate('agent1', 'Pass@123')
    f = _write_xlsx(os.path.join(d, "x.xlsx"), _rows(1))
    ok, result = imp.import_file(f)
    check("an agent cannot import history", not ok, result)
    check("nothing was written",
          dbm.execute_with_retry("SELECT COUNT(*) FROM reports")[0][0] == 0)


def test_import_is_refused_on_a_client():
    """A client holds a read-only replica. Refuse plainly rather than failing
    on 'attempt to write a readonly database'."""
    d, dbm, auth, reports, nums, imp = _setup()
    auth.authenticate('admin', 'Admin@1234')
    f = _write_xlsx(os.path.join(d, "c.xlsx"), _rows(1))
    dbm.read_only = True                      # as client mode builds it
    try:
        ok, result = imp.import_file(f)
        check("importing from a client is refused", not ok, result)
        check("and the message says where to run it",
              'host' in str(result.get('error', '')).lower(), result)
    finally:
        dbm.read_only = False


# ------------------------------------------------------ how they behave after
def test_archived_records_are_searchable_and_counted():
    d, dbm, auth, reports, nums, imp = _setup()
    auth.authenticate('admin', 'Admin@1234')
    imp.import_file(_write_xlsx(os.path.join(d, "h.xlsx"), _rows(3)))

    rows, total = reports.get_reports(None, None, None, None, None, 50, 0)
    check("archived reports show in the reports list", total == 3, total)
    rows2, t2 = reports.get_reports(None, 'Historic Entity 2', None, None, None, 50, 0)
    check("they are searchable by entity name", t2 >= 1, t2)
    rows3, t3 = reports.get_reports('archived', None, None, None, None, 50, 0)
    check("they can be filtered as archived", t3 == 3, t3)


def test_history_shows_in_the_cic_lookup_but_never_blocks():
    """The payoff: 20 years of customer history at the moment of writing."""
    d, dbm, auth, reports, nums, imp = _setup()
    from services.intelligence_service import IntelligenceService
    auth.authenticate('admin', 'Admin@1234')
    cic = '9100000000000001'
    imp.import_file(_write_xlsx(os.path.join(d, "h.xlsx"), _rows(1)))

    intel = IntelligenceService(dbm, None)
    hist = intel.cic_history(cic)
    check("an archived filing appears in the CIC history", hist['count'] == 1, hist['count'])
    prof = intel.customer_profile(cic)
    check("and the customer's details can be pulled from it",
          (prof or {}).get('reported_entity_name') == 'Historic Entity 1', prof)

    # a NEW report for that same customer must still be possible
    auth.authenticate('agent1', 'Pass@123')
    nums.reserve_block('agent1', 2)
    ok, rid, msg = reports.create_report({
        'report_date': '19/07/2026', 'reported_entity_name': 'Historic Entity 1 again',
        'cic': cic})
    check("a 2011 record never blocks a new report for that customer", ok, msg)


def test_archived_records_are_editable_by_anyone_who_may_edit():
    d, dbm, auth, reports, nums, imp = _setup()
    auth.authenticate('admin', 'Admin@1234')
    imp.import_file(_write_xlsx(os.path.join(d, "h.xlsx"), _rows(1)))
    rid = dbm.execute_with_retry("SELECT report_id FROM reports")[0][0]

    auth.authenticate('agent1', 'Pass@123')
    ok, msg = reports.update_report(rid, {'reported_entity_name': 'Corrected Name'})
    check("an agent can correct a transcription error", ok, msg)
    check("the correction stuck",
          (reports.get_report(rid) or {}).get('reported_entity_name') == 'Corrected Name')
    check("it is still archived after the edit",
          (reports.get_report(rid) or {}).get('approval_status') == 'archived')
    hist = reports.get_report_history(rid) or []
    check("and the correction is versioned from that point", len(hist) >= 1, len(hist))


def test_archived_records_stay_out_of_the_approval_queue():
    d, dbm, auth, reports, nums, imp = _setup()
    from services.approval_service import ApprovalService
    from services.version_service import VersionService
    from services.activity_service import ActivityService
    from services.logging_service import LoggingService
    auth.authenticate('admin', 'Admin@1234')
    imp.import_file(_write_xlsx(os.path.join(d, "h.xlsx"), _rows(2)))

    log = LoggingService(dbm, None, db_logging=False)
    act = ActivityService(dbm, log, auth)
    vers = VersionService(dbm, log, auth, reports, act)
    appr = ApprovalService(dbm, log, auth, vers, reports, act)
    auth.authenticate('sup1', 'Pass@123')
    pend = appr.get_pending_approvals() or []
    check("nothing archived sits in a supervisor's queue", not pend, pend)
    n = dbm.execute_with_retry("SELECT COUNT(*) FROM report_approvals")[0][0]
    check("no approval rows were invented", n == 0, n)


def test_provenance_is_recorded():
    d, dbm, auth, reports, nums, imp = _setup()
    auth.authenticate('admin', 'Admin@1234')
    f = _write_xlsx(os.path.join(d, "batch1.xlsx"), _rows(2))
    ok, result = imp.import_file(f)
    batch = result.get('batch_id')
    check("the import reports a batch id", bool(batch), result)
    rows = dbm.execute_with_retry(
        "SELECT DISTINCT import_batch_id FROM reports WHERE import_batch_id IS NOT NULL")
    check("every imported row carries it", len(rows) == 1 and rows[0][0] == batch, rows)
    b = dbm.execute_with_retry(
        "SELECT source_file, row_count, imported_by FROM import_batches WHERE batch_id = ?",
        (batch,))
    check("the batch records the source file and who ran it", bool(b), b)
    check("  with the row count", b and b[0][1] == 2, b)
    check("  and the file name", b and 'batch1' in (b[0][0] or ''), b)


if __name__ == "__main__":
    test_template_is_generated_by_the_app()
    test_clean_file_imports()
    test_numbers_are_registered_as_used()
    test_whole_file_is_rejected_and_nothing_is_written()
    test_every_problem_is_reported_in_one_pass()
    test_short_cics_are_padded_not_rejected()
    test_duplicate_numbers_are_caught_inside_the_file()
    test_a_number_already_in_str_is_caught()
    test_only_admin_may_import()
    test_import_is_refused_on_a_client()
    test_archived_records_are_searchable_and_counted()
    test_history_shows_in_the_cic_lookup_but_never_blocks()
    test_archived_records_are_editable_by_anyone_who_may_edit()
    test_archived_records_stay_out_of_the_approval_queue()
    test_provenance_is_recorded()
    print(f"\n{'ALL PASS' if _fail == 0 else str(_fail) + ' FAILED'}")
    sys.exit(1 if _fail else 0)
