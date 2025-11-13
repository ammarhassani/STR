"""
Reports view widget for managing financial crime reports.
Enhanced with advanced filtering, Excel export, and pagination.
"""

from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                             QPushButton, QTableWidget, QTableWidgetItem,
                             QLineEdit, QComboBox, QHeaderView, QMessageBox,
                             QFrame, QDateEdit, QSpinBox, QCheckBox, QFileDialog)
from PyQt6.QtCore import Qt, pyqtSignal, QDate
from PyQt6.QtGui import QFont, QColor
from ui.workers import ReportLoadWorker
from services.icon_service import get_icon
from pathlib import Path
from datetime import datetime


class ReportsView(QWidget):
    """
    Reports management view.

    Features:
    - View reports list
    - Search and filter
    - View/Edit reports
    """

    def __init__(self, report_service, logging_service, auth_service):
        """
        Initialize reports view.

        Args:
            report_service: ReportService instance
            logging_service: LoggingService instance
            auth_service: AuthService instance
        """
        super().__init__()
        self.report_service = report_service
        self.logging_service = logging_service
        self.auth_service = auth_service
        self.current_user = auth_service.get_current_user()
        self.current_reports = []
        self.worker = None

        # Pagination
        self.current_page = 1
        self.page_size = 50
        self.total_count = 0

        # Advanced filter state
        self.advanced_filters_visible = False

        self.setup_ui()
        self.load_reports()

    def setup_ui(self):
        """Setup the user interface."""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(16)

        # Header
        header_layout = QHBoxLayout()

        title_label = QLabel("Reports Management")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setWeight(QFont.Weight.Bold)
        title_label.setFont(title_font)
        header_layout.addWidget(title_label)

        header_layout.addStretch()

        # My Reports quick filter (for agents)
        if self.current_user.get('role') != 'admin':
            my_reports_btn = QPushButton("My Reports")
            my_reports_btn.setIcon(get_icon('user'))
            my_reports_btn.setObjectName("secondaryButton")
            my_reports_btn.setCheckable(True)
            my_reports_btn.clicked.connect(self.toggle_my_reports)
            my_reports_btn.setToolTip("Show only reports created by me")
            header_layout.addWidget(my_reports_btn)
            self.my_reports_btn = my_reports_btn

            # Send All to Approval button
            send_all_btn = QPushButton("Send All to Approval")
            send_all_btn.setIcon(get_icon('check-circle'))
            send_all_btn.setObjectName("primaryButton")
            send_all_btn.clicked.connect(self.send_all_to_approval)
            send_all_btn.setToolTip("Send all my draft reports for approval")
            header_layout.addWidget(send_all_btn)
            self.send_all_btn = send_all_btn

        # Export to Excel button
        export_btn = QPushButton("Export to Excel")
        export_btn.setIcon(get_icon('file-excel', color='#27ae60'))
        export_btn.setObjectName("secondaryButton")
        export_btn.clicked.connect(self.export_to_excel)
        header_layout.addWidget(export_btn)

        # Add Report button
        add_btn = QPushButton("Add New Report")
        add_btn.setIcon(get_icon('plus'))
        add_btn.setObjectName("primaryButton")
        add_btn.clicked.connect(self.add_report)
        header_layout.addWidget(add_btn)

        layout.addLayout(header_layout)

        # Basic Filters
        filters_layout = QHBoxLayout()

        # Search
        search_label = QLabel("Search:")
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search by report number, entity name, or CIC...")
        self.search_input.returnPressed.connect(self.on_filter_changed)
        filters_layout.addWidget(search_label)
        filters_layout.addWidget(self.search_input, stretch=1)

        # Status filter
        status_label = QLabel("Status:")
        self.status_combo = QComboBox()
        self.status_combo.addItems([
            'All', 'Open', 'Case Review', 'Under Investigation',
            'Case Validation', 'Close Case', 'Closed with STR'
        ])
        self.status_combo.currentTextChanged.connect(self.on_filter_changed)
        filters_layout.addWidget(status_label)
        filters_layout.addWidget(self.status_combo)

        # Advanced filters toggle
        self.advanced_btn = QPushButton("Advanced Filters")
        self.advanced_btn.setIcon(get_icon('filter'))
        self.advanced_btn.setCheckable(True)
        self.advanced_btn.clicked.connect(self.toggle_advanced_filters)
        filters_layout.addWidget(self.advanced_btn)

        # Clear filters
        clear_btn = QPushButton("Clear")
        clear_btn.setIcon(get_icon('times'))
        clear_btn.clicked.connect(self.clear_filters)
        filters_layout.addWidget(clear_btn)

        # Search/Refresh button
        search_btn = QPushButton("Search")
        search_btn.setIcon(get_icon('search'))
        search_btn.clicked.connect(self.on_filter_changed)
        filters_layout.addWidget(search_btn)

        layout.addLayout(filters_layout)

        # Advanced Filters Panel (collapsible)
        self.advanced_panel = QFrame()
        self.advanced_panel.setObjectName("card")
        self.advanced_panel.setVisible(False)
        advanced_layout = QVBoxLayout(self.advanced_panel)
        advanced_layout.setContentsMargins(16, 16, 16, 16)
        advanced_layout.setSpacing(12)

        advanced_title = QLabel("Advanced Filters")
        advanced_title.setStyleSheet("font-weight: 600; font-size: 11pt;")
        advanced_layout.addWidget(advanced_title)

        # Row 1: Date range
        date_row = QHBoxLayout()

        date_from_label = QLabel("Date From:")
        self.date_from_edit = QDateEdit()
        self.date_from_edit.setCalendarPopup(True)
        self.date_from_edit.setDate(QDate.currentDate().addMonths(-6))
        self.date_from_edit.setDisplayFormat("yyyy-MM-dd")
        date_row.addWidget(date_from_label)
        date_row.addWidget(self.date_from_edit)

        date_to_label = QLabel("Date To:")
        self.date_to_edit = QDateEdit()
        self.date_to_edit.setCalendarPopup(True)
        self.date_to_edit.setDate(QDate.currentDate())
        self.date_to_edit.setDisplayFormat("yyyy-MM-dd")
        date_row.addWidget(date_to_label)
        date_row.addWidget(self.date_to_edit)

        # Enable/disable date filter
        self.date_filter_check = QCheckBox("Enable Date Filter")
        date_row.addWidget(self.date_filter_check)
        date_row.addStretch()

        advanced_layout.addLayout(date_row)

        # Row 2: Created by
        creator_row = QHBoxLayout()

        creator_label = QLabel("Created By:")
        self.creator_combo = QComboBox()
        self.creator_combo.setMinimumWidth(200)
        creator_row.addWidget(creator_label)
        creator_row.addWidget(self.creator_combo)
        creator_row.addStretch()

        advanced_layout.addLayout(creator_row)

        layout.addWidget(self.advanced_panel)

        # Load creators for filter
        self.load_creators()

        # Stats and pagination
        stats_row = QHBoxLayout()
        self.stats_label = QLabel("0 reports")
        self.stats_label.setObjectName("subtitleLabel")
        stats_row.addWidget(self.stats_label)
        stats_row.addStretch()

        # Pagination controls
        pagination_widget = QFrame()
        pagination_layout = QHBoxLayout(pagination_widget)
        pagination_layout.setContentsMargins(0, 0, 0, 0)
        pagination_layout.setSpacing(8)

        self.prev_btn = QPushButton("Previous")
        self.prev_btn.setIcon(get_icon('chevron-left'))
        self.prev_btn.clicked.connect(self.previous_page)
        self.prev_btn.setEnabled(False)
        pagination_layout.addWidget(self.prev_btn)

        self.page_label = QLabel("Page 1")
        self.page_label.setStyleSheet("font-weight: 500; padding: 0 12px;")
        pagination_layout.addWidget(self.page_label)

        self.next_btn = QPushButton("Next")
        self.next_btn.setIcon(get_icon('chevron-right'))
        self.next_btn.clicked.connect(self.next_page)
        pagination_layout.addWidget(self.next_btn)

        page_size_label = QLabel("Per Page:")
        pagination_layout.addWidget(page_size_label)

        self.page_size_spin = QSpinBox()
        self.page_size_spin.setRange(10, 500)
        self.page_size_spin.setSingleStep(10)
        self.page_size_spin.setValue(self.page_size)
        self.page_size_spin.valueChanged.connect(self.on_page_size_changed)
        pagination_layout.addWidget(self.page_size_spin)

        stats_row.addWidget(pagination_widget)
        layout.addLayout(stats_row)

        # Reports table
        self.reports_table = QTableWidget()
        self.reports_table.setColumnCount(9)
        self.reports_table.setHorizontalHeaderLabels([
            'SN', 'Report Number', 'Date', 'Entity Name',
            'Status', 'Version', 'Approval', 'Created By', 'Created At'
        ])

        # Configure table
        self.reports_table.setAlternatingRowColors(True)
        self.reports_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.reports_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.reports_table.doubleClicked.connect(self.view_report)

        # Enable manual column resizing (drag column borders to resize)
        header = self.reports_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)  # All columns manually resizable
        header.setStretchLastSection(False)

        # Set default column widths
        header.resizeSection(0, 80)   # SN
        header.resizeSection(1, 150)  # Report Number
        header.resizeSection(2, 120)  # Date
        header.resizeSection(3, 250)  # Entity Name
        header.resizeSection(4, 150)  # Status
        header.resizeSection(5, 80)   # Version
        header.resizeSection(6, 120)  # Approval
        header.resizeSection(7, 130)  # Created By
        header.resizeSection(8, 170)  # Created At

        # Enable manual row resizing like Excel (drag row borders to resize)
        self.reports_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.reports_table.verticalHeader().setDefaultSectionSize(45)  # Default height for readability
        self.reports_table.verticalHeader().setMinimumSectionSize(30)  # Minimum to prevent too small

        # Connect signals to save geometry when user resizes
        header.sectionResized.connect(self.save_table_geometry)
        self.reports_table.verticalHeader().sectionResized.connect(self.save_table_geometry)

        layout.addWidget(self.reports_table)

        # Restore saved column widths and row heights
        self.restore_table_geometry()

        # Status label
        self.status_label = QLabel("")
        self.status_label.setObjectName("hintLabel")
        layout.addWidget(self.status_label)

    def toggle_advanced_filters(self):
        """Toggle advanced filters panel visibility."""
        self.advanced_filters_visible = not self.advanced_filters_visible
        self.advanced_panel.setVisible(self.advanced_filters_visible)

    def clear_filters(self):
        """Clear all filters and reset to defaults."""
        self.search_input.clear()
        self.status_combo.setCurrentIndex(0)
        self.date_filter_check.setChecked(False)
        self.date_from_edit.setDate(QDate.currentDate().addMonths(-6))
        self.date_to_edit.setDate(QDate.currentDate())
        self.creator_combo.setCurrentIndex(0)
        self.current_page = 1
        self.on_filter_changed()

    def load_creators(self):
        """Load list of users for creator filter."""
        try:
            # Get all users from database
            query = "SELECT user_id, username, full_name FROM users WHERE is_active = 1 ORDER BY full_name"
            users = self.report_service.db_manager.execute_with_retry(query)

            self.creator_combo.clear()
            self.creator_combo.addItem("All Users", None)

            for user in users:
                user_id, username, full_name = user
                display_name = f"{full_name} ({username})"
                # Store username (not user_id) because reports.created_by contains username
                self.creator_combo.addItem(display_name, username)

        except Exception as e:
            self.logging_service.error(f"Error loading creators: {str(e)}")

    def previous_page(self):
        """Go to previous page."""
        if self.current_page > 1:
            self.current_page -= 1
            self.load_reports()

    def next_page(self):
        """Go to next page."""
        if self.current_page * self.page_size < self.total_count:
            self.current_page += 1
            self.load_reports()

    def on_page_size_changed(self, value: int):
        """Handle page size change."""
        self.page_size = value
        self.current_page = 1
        self.load_reports()

    def on_filter_changed(self):
        """Handle filter change."""
        self.current_page = 1
        self.load_reports()

    def load_reports(self):
        """Load reports from database with filters and pagination."""
        self.status_label.setText("Loading reports...")

        # Get filter values
        status = None if self.status_combo.currentText() == 'All' else self.status_combo.currentText()
        search_term = self.search_input.text().strip() or None

        # Advanced filters
        date_from = None
        date_to = None
        if self.date_filter_check.isChecked():
            date_from = self.date_from_edit.date().toString("yyyy-MM-dd")
            date_to = self.date_to_edit.date().toString("yyyy-MM-dd")

        created_by = self.creator_combo.currentData()

        # Calculate offset for pagination
        offset = (self.current_page - 1) * self.page_size

        # Load in worker thread
        self.worker = ReportLoadWorker(
            self.report_service,
            status=status,
            search_term=search_term,
            date_from=date_from,
            date_to=date_to,
            created_by=created_by,
            limit=self.page_size,
            offset=offset
        )
        self.worker.finished.connect(self.on_reports_loaded)
        self.worker.error.connect(self.on_load_error)
        self.worker.progress.connect(lambda v, m: self.status_label.setText(m))
        self.worker.start()

    def on_reports_loaded(self, reports: list, total_count: int):
        """
        Handle reports loaded.

        Args:
            reports: List of report dictionaries
            total_count: Total number of reports matching filter
        """
        self.current_reports = reports
        self.total_count = total_count

        # Update table
        self.reports_table.setRowCount(0)
        self.reports_table.setRowCount(len(reports))

        for row, report in enumerate(reports):
            # SN
            sn_item = QTableWidgetItem(str(report.get('sn', '')))
            self.reports_table.setItem(row, 0, sn_item)

            # Report Number
            report_num_item = QTableWidgetItem(report.get('report_number', ''))
            self.reports_table.setItem(row, 1, report_num_item)

            # Date
            date_item = QTableWidgetItem(report.get('report_date', ''))
            self.reports_table.setItem(row, 2, date_item)

            # Entity Name
            entity_item = QTableWidgetItem(report.get('reported_entity_name', ''))
            self.reports_table.setItem(row, 3, entity_item)

            # Status
            status_item = QTableWidgetItem(report.get('status', ''))
            self.reports_table.setItem(row, 4, status_item)

            # Version
            version = report.get('current_version', 1)
            version_item = QTableWidgetItem(f"v{version}")
            version_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.reports_table.setItem(row, 5, version_item)

            # Approval Status
            approval_status = report.get('approval_status', 'draft')
            approval_labels = {
                'draft': 'Draft',
                'pending_approval': 'Pending',
                'approved': 'Approved',
                'rejected': 'Rejected',
                'rework': 'Rework'
            }
            approval_colors = {
                'draft': '#6e7681',
                'pending_approval': '#d29922',
                'approved': '#2ea043',
                'rejected': '#f85149',
                'rework': '#d29922'
            }
            approval_label = approval_labels.get(approval_status, approval_status)
            approval_item = QTableWidgetItem(approval_label)
            approval_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            # Color-code approval status
            approval_color = approval_colors.get(approval_status, '#6e7681')
            approval_item.setForeground(QColor(approval_color))

            self.reports_table.setItem(row, 6, approval_item)

            # Created By
            created_by_item = QTableWidgetItem(report.get('created_by', ''))
            self.reports_table.setItem(row, 7, created_by_item)

            # Created At
            created_at = report.get('created_at', '')
            if created_at:
                created_at = created_at[:19]  # Remove microseconds
            created_at_item = QTableWidgetItem(created_at)
            self.reports_table.setItem(row, 8, created_at_item)

        # Update stats
        start_record = (self.current_page - 1) * self.page_size + 1
        end_record = min(start_record + len(reports) - 1, total_count)
        self.stats_label.setText(f"Showing {start_record}-{end_record} of {total_count} reports")
        self.status_label.setText("Reports loaded")

        # Update pagination controls
        total_pages = (total_count + self.page_size - 1) // self.page_size
        self.page_label.setText(f"Page {self.current_page} of {total_pages}")

        self.prev_btn.setEnabled(self.current_page > 1)
        self.next_btn.setEnabled(self.current_page < total_pages)

    def on_load_error(self, error_message: str):
        """
        Handle load error.

        Args:
            error_message: Error message
        """
        self.status_label.setText(f"Error: {error_message}")
        QMessageBox.critical(self, "Error", f"Failed to load reports:\n{error_message}")

    def view_report(self):
        """View selected report."""
        selected_rows = self.reports_table.selectedIndexes()
        if not selected_rows:
            return

        row = selected_rows[0].row()
        if row < len(self.current_reports):
            report = self.current_reports[row]

            # Import here to avoid circular imports
            from ui.dialogs.report_dialog import ReportDialog

            dialog = ReportDialog(
                self.report_service,
                self.logging_service,
                self.current_user,
                report_data=report,
                parent=self
            )
            dialog.report_saved.connect(self.load_reports)
            dialog.exec()

    def add_report(self):
        """Add new report."""
        # Import here to avoid circular imports
        from ui.dialogs.report_dialog import ReportDialog

        dialog = ReportDialog(
            self.report_service,
            self.logging_service,
            self.current_user,
            parent=self
        )
        dialog.report_saved.connect(self.load_reports)
        dialog.exec()

    def export_to_excel(self):
        """Export filtered reports to Excel."""
        try:
            # Check if there are reports to export
            if not self.current_reports:
                QMessageBox.warning(
                    self,
                    "No Data",
                    "There are no reports to export. Please load some reports first."
                )
                return

            # Get save file name
            default_filename = f"FIU_Reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Export to Excel",
                str(Path.home() / "Downloads" / default_filename),
                "Excel Files (*.xlsx);;All Files (*)"
            )

            if not file_path:
                return  # User cancelled

            # Import openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from openpyxl.utils import get_column_letter
            except ImportError:
                QMessageBox.critical(
                    self,
                    "Missing Dependency",
                    "The openpyxl library is required for Excel export.\n"
                    "Please install it using: pip install openpyxl"
                )
                return

            # Show progress
            self.status_label.setText("Exporting to Excel...")

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "FIU Reports"

            # Define styles
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="0d7377", end_color="0d7377", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Headers
            headers = [
                'SN', 'Report Number', 'Report Date', 'Entity Name', 'CIC Number',
                'Transaction Amount', 'Status', 'Nature of Report', 'Action Taken',
                'Created By', 'Created At', 'Updated At'
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Fetch ALL matching reports (not just current page)
            status = None if self.status_combo.currentText() == 'All' else self.status_combo.currentText()
            search_term = self.search_input.text().strip() or None

            date_from = None
            date_to = None
            if self.date_filter_check.isChecked():
                date_from = self.date_from_edit.date().toString("yyyy-MM-dd")
                date_to = self.date_to_edit.date().toString("yyyy-MM-dd")

            created_by = self.creator_combo.currentData()

            # Get all reports for export (without pagination limit)
            all_reports, _ = self.report_service.get_reports(
                status=status,
                search_term=search_term,
                date_from=date_from,
                date_to=date_to,
                created_by=created_by,
                limit=None  # Get all
            )

            # Data rows
            for row, report in enumerate(all_reports, 2):
                ws.cell(row=row, column=1, value=report.get('sn', '')).border = border
                ws.cell(row=row, column=2, value=report.get('report_number', '')).border = border
                ws.cell(row=row, column=3, value=report.get('report_date', '')).border = border
                ws.cell(row=row, column=4, value=report.get('reported_entity_name', '')).border = border
                ws.cell(row=row, column=5, value=report.get('cic_number', '')).border = border
                ws.cell(row=row, column=6, value=report.get('transaction_amount', '')).border = border
                ws.cell(row=row, column=7, value=report.get('status', '')).border = border
                ws.cell(row=row, column=8, value=report.get('nature_of_report', '')).border = border
                ws.cell(row=row, column=9, value=report.get('action_taken', '')).border = border
                ws.cell(row=row, column=10, value=report.get('created_by', '')).border = border

                created_at = report.get('created_at', '')
                if created_at:
                    created_at = created_at[:19]
                ws.cell(row=row, column=11, value=created_at).border = border

                updated_at = report.get('updated_at', '')
                if updated_at:
                    updated_at = updated_at[:19]
                ws.cell(row=row, column=12, value=updated_at).border = border

            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                max_length = len(headers[col - 1])

                for cell in ws[column_letter]:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Freeze header row
            ws.freeze_panes = "A2"

            # Save workbook
            wb.save(file_path)

            self.status_label.setText(f"Exported {len(all_reports)} reports to Excel")
            self.logging_service.info(f"Exported {len(all_reports)} reports to {file_path}")

            # Show success message
            reply = QMessageBox.question(
                self,
                "Export Successful",
                f"Successfully exported {len(all_reports)} reports to:\n{file_path}\n\n"
                "Would you like to open the file?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

            if reply == QMessageBox.StandardButton.Yes:
                import os
                os.startfile(file_path)  # Windows-specific

        except Exception as e:
            error_msg = f"Failed to export to Excel: {str(e)}"
            self.status_label.setText("Export failed")
            self.logging_service.error(error_msg)
            QMessageBox.critical(self, "Export Error", error_msg)

    def toggle_my_reports(self):
        """Toggle filter to show only current user's reports."""
        if hasattr(self, 'my_reports_btn') and self.my_reports_btn.isChecked():
            # Set creator filter to current user
            for i in range(self.creator_combo.count()):
                if self.creator_combo.itemData(i) == self.current_user['username']:
                    self.creator_combo.setCurrentIndex(i)
                    break
            self.my_reports_btn.setText("All Reports")
            self.my_reports_btn.setToolTip("Show all reports")
        else:
            # Clear creator filter
            self.creator_combo.setCurrentIndex(0)  # "All Users"
            if hasattr(self, 'my_reports_btn'):
                self.my_reports_btn.setText("My Reports")
                self.my_reports_btn.setToolTip("Show only reports created by me")

        # Reload reports with new filter
        self.on_filter_changed()

    def send_all_to_approval(self):
        """Send all draft reports created by current user to approval."""
        try:
            # Get all draft reports by current user
            draft_reports, _ = self.report_service.get_reports(
                status=None,
                search_term=None,
                created_by=self.current_user['username'],
                limit=None  # Get all
            )

            # Filter for draft approval status
            draft_approval_reports = [
                r for r in draft_reports
                if r.get('approval_status', 'draft') == 'draft'
            ]

            if not draft_approval_reports:
                QMessageBox.information(
                    self,
                    "No Draft Reports",
                    "You have no draft reports to send for approval."
                )
                return

            # Confirm action
            reply = QMessageBox.question(
                self,
                "Confirm Send All",
                f"Send {len(draft_approval_reports)} draft report(s) for approval?\n\n"
                "This will submit all your draft reports for admin review.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

            if reply != QMessageBox.StandardButton.Yes:
                return

            # Send each report for approval
            success_count = 0
            failed_reports = []

            for report in draft_approval_reports:
                success, approval_id, message = self.report_service.request_approval(
                    report['report_id'],
                    comment="Bulk submission via Send All"
                )

                if success:
                    success_count += 1
                else:
                    failed_reports.append(f"Report {report.get('report_number', report['report_id'])}: {message}")

            # Show results
            if success_count == len(draft_approval_reports):
                QMessageBox.information(
                    self,
                    "Success",
                    f"Successfully sent {success_count} report(s) for approval!"
                )
            elif success_count > 0:
                QMessageBox.warning(
                    self,
                    "Partial Success",
                    f"Sent {success_count} of {len(draft_approval_reports)} reports for approval.\n\n"
                    f"Failed reports:\n" + "\n".join(failed_reports[:5])
                )
            else:
                QMessageBox.critical(
                    self,
                    "Failed",
                    f"Failed to send reports for approval:\n" + "\n".join(failed_reports[:5])
                )

            # Reload reports to show updated approval status
            self.load_reports()

        except Exception as e:
            error_msg = f"Error sending reports for approval: {str(e)}"
            self.logging_service.error(error_msg, exc_info=True)
            QMessageBox.critical(self, "Error", error_msg)

    def refresh(self):
        """Refresh the view (called from main window)."""
        self.load_reports()

    def save_table_geometry(self):
        """Save column widths and row heights to settings."""
        from PyQt6.QtCore import QSettings
        settings = QSettings('FIU', 'ReportManagement')

        # Save column widths
        column_widths = []
        for i in range(self.reports_table.columnCount()):
            column_widths.append(self.reports_table.columnWidth(i))
        settings.setValue('reports_view/column_widths', column_widths)

        # Save default row height (when user resizes any row, apply to all)
        if self.reports_table.rowCount() > 0:
            # Get the height of the first row as the default for all rows
            default_height = self.reports_table.rowHeight(0)
            settings.setValue('reports_view/default_row_height', default_height)

    def restore_table_geometry(self):
        """Restore column widths and row heights from settings."""
        from PyQt6.QtCore import QSettings
        settings = QSettings('FIU', 'ReportManagement')

        # Restore column widths
        column_widths = settings.value('reports_view/column_widths', None)
        if column_widths:
            for i, width in enumerate(column_widths):
                if i < self.reports_table.columnCount():
                    self.reports_table.setColumnWidth(i, int(width))

        # Restore default row height
        default_height = settings.value('reports_view/default_row_height', None)
        if default_height:
            self.reports_table.verticalHeader().setDefaultSectionSize(int(default_height))
