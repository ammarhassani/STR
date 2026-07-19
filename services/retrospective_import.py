"""Retrospective import: the unit's pre-STR history, from Excel into the app.

Design: docs/superpowers/specs/2026-07-19-retrospective-import-design.md

The unit filed reports for ~20 years before this app existed, in spreadsheets.
Those files are cleansed outside STR and then brought in through here, so the
app holds one continuous record: "imagine the FIU app existed from 2005 until
now."

Three things make this different from ordinary report creation:

  * the original report number is kept EXACTLY -- it is what was filed with the
    FIU and what an auditor searches for -- and is registered as used, so the
    live sequence can never hand it out again;
  * the records land as `archived`, never `approved`: they did not go through
    this app's workflow, so there is no approval row, no version history and no
    invented filer;
  * the file is all-or-nothing. Nothing is written unless every row is clean,
    which is why validation reports EVERY problem in one pass -- with 60,000
    rows, one error per upload would never converge.
"""
import os
import uuid
from datetime import datetime
from typing import Dict, List, Optional, Tuple

# The status an imported record carries. Deliberately outside the workflow
# statuses so nothing in the approval machinery can pick one up.
ARCHIVED_STATUS = 'archived'


def _clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    return str(value).strip()


class RetrospectiveImportService:
    """Generates the template, validates a filled one, and imports it."""

    # The columns the importer reads. The template is generated from this list,
    # so the file always originates from the app and the header can never drift
    # away from what the code expects.
    TEMPLATE_COLUMNS = (
        'report_number', 'sn', 'report_date', 'reported_entity_name',
        'cic', 'customer_segment', 'gender', 'nationality', 'id_cr', 'id_type',
        'account_membership', 'branch_id', 'relationship',
        'first_reason_for_suspicion', 'second_reason_for_suspicion',
        'type_of_suspected_transaction', 'total_transaction',
        'report_classification', 'report_source', 'reporting_entity',
        'arb_staff', 'outgoing_letter_number', 'sending_date',
        'fiu_number', 'fiu_date', 'fiu_letter_number', 'fiu_letter_receive_date',
        'fiu_feedback', 'case_id',
    )

    # Without these four a record cannot be identified or found again. The FIU
    # fields are NOT here: they gate submission for a live report, and an
    # archived record is never submitted.
    REQUIRED_COLUMNS = ('report_number', 'sn', 'report_date', 'reported_entity_name')

    MAX_EXAMPLE_ROWS = 20      # per problem type, in the summary

    def __init__(self, db_manager, logging_service, auth_service,
                 report_service=None, report_number_service=None):
        self.db_manager = db_manager
        self.logger = logging_service
        self.auth_service = auth_service
        self.report_service = report_service
        self.report_number_service = report_number_service

    # ------------------------------------------------------------- authority
    def _require_admin(self, action: str) -> Tuple[bool, str]:
        user = (self.auth_service.get_current_user() if self.auth_service else None) or {}
        if not user:
            return False, "Not authenticated"
        if user.get('role') != 'admin':
            if self.logger:
                self.logger.warning(
                    f"Unauthorized retrospective {action} attempt by "
                    f"{user.get('username')} (role={user.get('role')})")
            return False, "Administrator privileges required"
        # A client works against a read-only replica. Importing 60,000 rows
        # through the command queue would be absurd, so this runs on the host
        # PC -- say so plainly instead of failing on a read-only database.
        if action == "import" and getattr(self.db_manager, "read_only", False):
            return False, ("Historical import must be run on the host PC. This "
                           "workstation is a client and holds a read-only copy.")
        return True, ""

    # -------------------------------------------------------------- template
    def write_template(self, path: str) -> Tuple[bool, str]:
        """Write the blank template the historical data is pasted into."""
        allowed, why = self._require_admin("template")
        if not allowed:
            return False, why
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            wb = Workbook()
            ws = wb.active
            ws.title = "reports"
            ws.append(list(self.TEMPLATE_COLUMNS))
            for i, col in enumerate(self.TEMPLATE_COLUMNS, start=1):
                cell = ws.cell(row=1, column=i)
                cell.font = Font(bold=True)
                # a required column is marked so it is obvious before filling
                if col in self.REQUIRED_COLUMNS:
                    cell.value = col          # header must stay machine-readable
                ws.column_dimensions[cell.column_letter].width = max(14, len(col) + 2)
            ws.freeze_panes = "A2"

            notes = wb.create_sheet("how to use")
            for line in (
                "Retrospective import template",
                "",
                "One row per historical report. Do not rename, reorder or remove "
                "columns on the 'reports' sheet - the importer reads them by name.",
                "",
                "Required in every row: " + ", ".join(self.REQUIRED_COLUMNS),
                "report_number: exactly as filed, e.g. 2016/04/555",
                "sn: the serial number, a whole number, unique across all reports",
                "report_date: DD/MM/YYYY",
                "cic: 16 digits (leading zeros are fine)",
                "",
                "The whole file is refused unless every row is clean. The refusal "
                "lists every problem found, grouped by type.",
            ):
                notes.append([line])
            wb.save(path)
            return True, f"Template written to {path}"
        except Exception as e:
            if self.logger:
                self.logger.error(f"Could not write the import template: {e}", exc_info=True)
            return False, f"Could not write the template: {e}"

    # ------------------------------------------------------------ validation
    def _existing_keys(self) -> Tuple[set, set]:
        """Report numbers and serials already in STR.

        Loaded once. Checking 60,000 rows one query at a time would be 120,000
        round trips.
        """
        numbers = {r[0] for r in (self.db_manager.execute_with_retry(
            "SELECT report_number FROM reports") or []) if r[0]}
        numbers |= {r[0] for r in (self.db_manager.execute_with_retry(
            "SELECT report_number FROM reserved_numbers") or []) if r[0]}
        serials = {r[0] for r in (self.db_manager.execute_with_retry(
            "SELECT sn FROM reports") or []) if r[0] is not None}
        return numbers, serials

    def read_rows(self, path: str) -> List[Dict]:
        """Stream the filled template. read_only so 60k rows do not load whole."""
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb["reports"] if "reports" in wb.sheetnames else wb.worksheets[0]
        rows, header = [], None
        for raw in ws.iter_rows(values_only=True):
            if header is None:
                header = [_clean(c) for c in raw]
                continue
            if all(c is None or _clean(c) == "" for c in raw):
                continue                      # blank spacer row
            rows.append({h: raw[i] if i < len(raw) else None
                         for i, h in enumerate(header) if h})
        wb.close()
        return rows

    def validate(self, rows: List[Dict]) -> Dict:
        """Check every row and report every problem. Writes nothing.

        Returns {'bad_rows': n, 'problems': {kind: {count, rows}}, 'clean': [...]}
        """
        from services.report_service import _date_key

        existing_numbers, existing_serials = self._existing_keys()
        seen_numbers, seen_serials = set(), set()
        problems: Dict[str, Dict] = {}
        bad_row_numbers = set()
        clean: List[Dict] = []

        def problem(kind: str, excel_row: int):
            entry = problems.setdefault(kind, {"count": 0, "rows": []})
            entry["count"] += 1
            if len(entry["rows"]) < self.MAX_EXAMPLE_ROWS:
                entry["rows"].append(excel_row)
            bad_row_numbers.add(excel_row)

        for index, row in enumerate(rows):
            excel_row = index + 2            # header is row 1
            values = {k: _clean(v) for k, v in row.items()}

            for col in self.REQUIRED_COLUMNS:
                if not values.get(col):
                    problem(f"{col} is missing", excel_row)

            number = values.get('report_number', '')
            if number:
                if number in existing_numbers:
                    problem("report_number already exists in STR", excel_row)
                if number in seen_numbers:
                    problem("duplicate report_number inside this file", excel_row)
                seen_numbers.add(number)

            serial = values.get('sn', '')
            if serial:
                if not serial.isdigit():
                    problem("sn is not a whole number", excel_row)
                else:
                    as_int = int(serial)
                    if as_int in existing_serials:
                        problem("sn already exists in STR", excel_row)
                    if as_int in seen_serials:
                        problem("duplicate sn inside this file", excel_row)
                    seen_serials.add(as_int)

            for date_col in ('report_date', 'sending_date', 'fiu_date',
                             'fiu_letter_receive_date'):
                raw = values.get(date_col, '')
                if raw and _date_key(raw) is None:
                    problem(f"{date_col} is not a real date (DD/MM/YYYY)", excel_row)

            cic = values.get('cic', '')
            if cic:
                digits = cic.replace(' ', '').replace('-', '')
                if not digits.isdigit():
                    problem("cic contains something other than digits", excel_row)
                elif len(digits) > 16:
                    problem("cic is longer than 16 digits", excel_row)
                else:
                    # A CIC is a fixed 16-digit code, and the bank's own systems
                    # print it without leading zeros -- the report form pads it
                    # on entry for exactly this reason. Historical spreadsheets
                    # carry the short form, so pad rather than reject, or the
                    # imported CIC would never match one typed today.
                    values['cic'] = digits.zfill(16)

            amount = values.get('total_transaction', '')
            if amount:
                try:
                    if float(amount.replace(',', '')) < 0:
                        problem("total_transaction is negative", excel_row)
                except ValueError:
                    problem("total_transaction is not a number", excel_row)

            if excel_row not in bad_row_numbers:
                clean.append(values)

        return {"bad_rows": len(bad_row_numbers), "problems": problems,
                "clean": clean, "total_rows": len(rows)}

    # ---------------------------------------------------------------- import
    def import_file(self, path: str) -> Tuple[bool, Dict]:
        """Validate a filled template and, only if it is spotless, import it."""
        allowed, why = self._require_admin("import")
        if not allowed:
            return False, {"error": why}

        try:
            rows = self.read_rows(path)
        except Exception as e:
            return False, {"error": f"Could not read the file: {e}"}
        if not rows:
            return False, {"error": "The file has no data rows"}

        report = self.validate(rows)
        if report["bad_rows"]:
            report["error"] = (
                f"{report['bad_rows']} of {report['total_rows']} rows have problems. "
                f"Nothing was imported. Fix them and upload the file again.")
            if self.logger:
                self.logger.warning(
                    f"Retrospective import refused: {report['bad_rows']} bad rows in "
                    f"{os.path.basename(path)}")
            return False, report

        return self._insert(report["clean"], path)

    def _insert(self, rows: List[Dict], path: str) -> Tuple[bool, Dict]:
        """Write every row in ONE transaction, or none of them."""
        import sqlite3
        user = (self.auth_service.get_current_user() or {}).get('username', 'SYSTEM')
        batch_id = uuid.uuid4().hex
        now = datetime.now().isoformat()
        columns = [c for c in self.TEMPLATE_COLUMNS]

        conn = sqlite3.connect(self.db_manager.db_path)
        try:
            cur = conn.cursor()
            cur.execute("BEGIN IMMEDIATE")
            cur.execute(
                "INSERT INTO import_batches (batch_id, source_file, row_count, "
                "imported_by, imported_at) VALUES (?, ?, ?, ?, ?)",
                (batch_id, os.path.basename(path), len(rows), user, now))

            insert_cols = columns + ['approval_status', 'created_by', 'created_at',
                                     'import_batch_id', 'is_deleted']
            placeholders = ",".join("?" for _ in insert_cols)
            sql = (f"INSERT INTO reports ({','.join(insert_cols)}) "
                   f"VALUES ({placeholders})")

            for row in rows:
                values = [row.get(c) or None for c in columns]
                # sn is an integer column
                sn_index = columns.index('sn')
                values[sn_index] = int(row['sn'])
                cur.execute(sql, values + [ARCHIVED_STATUS, user, now, batch_id, 0])
                report_id = cur.lastrowid
                # register the number so the live sequence can never re-issue it
                cur.execute(
                    "INSERT INTO reserved_numbers (report_number, serial_number, "
                    "owned_by, status, used_by_report_id, reserved_at) "
                    "VALUES (?, ?, ?, 'used', ?, ?)",
                    (row['report_number'], int(row['sn']), user, report_id, now))

            conn.commit()
        except Exception as e:
            conn.rollback()
            if self.logger:
                self.logger.error(f"Retrospective import failed, rolled back: {e}",
                                  exc_info=True)
            return False, {"error": f"Import failed and nothing was written: {e}"}
        finally:
            conn.close()

        if self.logger:
            self.logger.info(
                f"Retrospective import: {len(rows)} historical reports from "
                f"{os.path.basename(path)} (batch {batch_id}) by {user}")
            try:
                self.logger.log_user_action(
                    "HISTORY_IMPORTED",
                    {"batch_id": batch_id, "rows": len(rows),
                     "source_file": os.path.basename(path), "imported_by": user})
            except Exception:
                pass
        return True, {"imported": len(rows), "batch_id": batch_id,
                      "source_file": os.path.basename(path)}

    # ------------------------------------------------------------- batch ops
    def list_batches(self) -> List[Dict]:
        rows = self.db_manager.execute_with_retry(
            "SELECT batch_id, source_file, row_count, imported_by, imported_at "
            "FROM import_batches ORDER BY imported_at DESC") or []
        return [{"batch_id": r[0], "source_file": r[1], "row_count": r[2],
                 "imported_by": r[3], "imported_at": r[4]} for r in rows]
